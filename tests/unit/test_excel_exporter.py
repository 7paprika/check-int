from openpyxl import load_workbook

from check_int.services.excel_exporter import export_comparison_rows_to_excel


def test_export_comparison_rows_to_excel_writes_rows_and_highlights_mismatch(tmp_path) -> None:
    output_path = tmp_path / "integrity-report.xlsx"
    rows = [
        {
            "tag_no": "P-601",
            "field_name": "service",
            "status": "matched",
            "master_value": "Cooling Water",
            "pid_value": "Cooling Water",
            "datasheet_value": "Cooling Water",
        },
        {
            "tag_no": "P-601",
            "field_name": "material",
            "status": "mismatch",
            "master_value": "SS316",
            "pid_value": "CS",
            "datasheet_value": "SS316",
        },
    ]

    export_comparison_rows_to_excel(rows, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["Integrity Report"]

    assert sheet["A1"].value == "Tag No"
    assert sheet["B1"].value == "Field"
    assert sheet["C1"].value == "Status"
    assert sheet["A2"].value == "P-601"
    assert sheet["B3"].value == "Material"
    assert sheet["C3"].value == "Mismatch"
    assert sheet["E3"].value == "CS"
    assert sheet["F3"].value == "SS316"
    assert sheet["E3"].fill.fill_type == "solid"
    assert sheet["E3"].fill.fgColor.rgb.endswith("FFF1B3")


def test_export_comparison_rows_to_excel_can_export_mismatch_only_with_summary(tmp_path) -> None:
    output_path = tmp_path / "integrity-report.xlsx"
    rows = [
        {
            "tag_no": "P-701",
            "field_name": "service",
            "status": "matched",
            "master_value": "Water",
            "pid_value": "Water",
            "datasheet_value": "Water",
        },
        {
            "tag_no": "P-702",
            "field_name": "material",
            "status": "mismatch",
            "master_value": "SS316",
            "pid_value": "CS",
            "datasheet_value": "SS316",
        },
        {
            "tag_no": "P-703",
            "field_name": "tag_no",
            "status": "duplicate_tag",
            "master_value": "P-703",
            "pid_value": "P-703 x2",
            "datasheet_value": "P-703",
        },
    ]

    export_comparison_rows_to_excel(rows, output_path, mismatch_only=True)

    workbook = load_workbook(output_path)
    detail = workbook["Integrity Report"]
    summary = workbook["Summary"]
    assert detail.max_row == 3
    assert detail["A2"].value == "P-702"
    assert detail["A3"].value == "P-703"
    assert summary["A1"].value == "Status"
    assert summary["B1"].value == "Count"
    summary_counts = {summary.cell(row=i, column=1).value: summary.cell(row=i, column=2).value for i in range(2, summary.max_row + 1)}
    assert summary_counts == {"Mismatch": 1, "Duplicate Tag": 1}
