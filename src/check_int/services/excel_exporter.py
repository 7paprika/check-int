from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import PatternFill


HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFF1B3")
HEADERS = [
    "Tag No",
    "Field",
    "Status",
    "Master",
    "P&ID",
    "Datasheet",
    "Evidence Note",
    "P&ID Evidence",
    "Datasheet Evidence",
]

FIELD_LABELS = {
    "tag_no": "Tag No",
    "service": "Service",
    "material": "Material",
    "capacity": "Capacity",
    "size": "Size",
    "model": "Model",
    "design_temperature": "Design Temperature",
    "design_pressure": "Design Pressure",
    "operating_pressure": "Operating Pressure",
    "operating_temperature": "Operating Temperature",
}

STATUS_LABELS = {
    "matched": "Matched",
    "mismatch": "Mismatch",
    "missing_target": "Missing Target",
    "missing_source": "Missing Source",
    "unreviewed": "Unreviewed",
    "duplicate_tag": "Duplicate Tag",
}


def export_comparison_rows_to_excel(
    rows: list[dict[str, object | None]],
    output_path: str | Path,
    *,
    embed_images: bool = False,
    mismatch_only: bool = False,
) -> Path:
    export_rows = _filter_rows(rows, mismatch_only=mismatch_only)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Integrity Report"
    _write_summary_sheet(workbook, export_rows)
    sheet.append(HEADERS)
    _apply_column_widths(sheet)

    for row in export_rows:
        sheet.append(
            [
                row.get("tag_no"),
                FIELD_LABELS.get(str(row.get("field_name") or ""), str(row.get("field_name") or "")),
                STATUS_LABELS.get(str(row.get("status") or ""), str(row.get("status") or "")),
                row.get("master_value"),
                row.get("pid_value"),
                row.get("datasheet_value"),
                _evidence_note(row),
                row.get("pid_image_path"),
                row.get("datasheet_image_path"),
            ]
        )
        current_row = sheet.max_row
        if row.get("status") in {"mismatch", "duplicate_tag"}:
            for column_index in range(4, 7):
                sheet.cell(row=current_row, column=column_index).fill = HIGHLIGHT_FILL
        if embed_images:
            _embed_row_images(sheet, row, current_row)

    output = Path(output_path)
    workbook.save(output)
    return output


def _filter_rows(rows: list[dict[str, object | None]], *, mismatch_only: bool) -> list[dict[str, object | None]]:
    if not mismatch_only:
        return list(rows)
    return [row for row in rows if row.get("status") != "matched"]


def _write_summary_sheet(workbook: Workbook, rows: list[dict[str, object | None]]) -> None:
    summary = workbook.create_sheet("Summary", 0)
    summary.append(["Status", "Count"])
    counts = Counter(str(row.get("status") or "") for row in rows)
    for status, count in sorted(counts.items()):
        summary.append([STATUS_LABELS.get(status, status), count])
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 12


def _apply_column_widths(sheet) -> None:
    widths = {
        "A": 16,
        "B": 24,
        "C": 18,
        "D": 20,
        "E": 20,
        "F": 20,
        "G": 30,
        "H": 18,
        "I": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _evidence_note(row: dict[str, object | None]) -> str | None:
    notes: list[str] = []
    if row.get("pid_page_no"):
        notes.append(f"P&ID p.{row['pid_page_no']}")
    if row.get("datasheet_page_no"):
        notes.append(f"Datasheet p.{row['datasheet_page_no']}")
    return "; ".join(notes) or None


def _embed_row_images(sheet, row: dict[str, object | None], current_row: int) -> None:
    pid_image_path = row.get("pid_image_path")
    datasheet_image_path = row.get("datasheet_image_path")
    embedded = False

    if pid_image_path and Path(str(pid_image_path)).exists():
        pid_image = OpenPyxlImage(str(pid_image_path))
        pid_image.width = 80
        pid_image.height = 80
        sheet.add_image(pid_image, f"H{current_row}")
        embedded = True

    if datasheet_image_path and Path(str(datasheet_image_path)).exists():
        datasheet_image = OpenPyxlImage(str(datasheet_image_path))
        datasheet_image.width = 80
        datasheet_image.height = 80
        sheet.add_image(datasheet_image, f"I{current_row}")
        embedded = True

    if embedded:
        sheet.row_dimensions[current_row].height = 60
